#!/usr/bin/env python3
"""Create a formatted Excel backup for the Golf Majors Sweepstake 2026."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

PLAYERS = ["Ross", "Euan", "AJ", "Ryan", "Sam", "Stuart"]
MAJORS = [
    {"name": "The Masters", "venue": "Augusta National GC", "date": "10-13 April 2026"},
    {"name": "PGA Championship", "venue": "Aronimink Golf Club", "date": "14-17 May 2026"},
    {"name": "US Open", "venue": "Shinnecock Hills GC", "date": "18-21 June 2026"},
    {"name": "The Open", "venue": "Royal Birkdale", "date": "16-19 July 2026"},
]
SLOTS = ["European 1", "European 2", "American", "Rest of World"]

# Colours
MASTERS_GREEN = "165F3B"
MASTERS_GOLD = "D4AF37"
WHITE = "FFFFFF"
LIGHT_GREY = "F7F7F7"
MID_GREY = "E5E5E5"
DARK_TEXT = "1A1A1A"
RED = "DC2626"
GREEN = "16A34A"
PGA_NAVY = "003366"
USO_RED = "C41E3A"
OPEN_NAVY = "1A1A3E"

MAJOR_COLOURS = {
    "The Masters": MASTERS_GREEN,
    "PGA Championship": PGA_NAVY,
    "US Open": USO_RED,
    "The Open": OPEN_NAVY,
}

thin_border = Border(
    left=Side(style="thin", color=MID_GREY),
    right=Side(style="thin", color=MID_GREY),
    top=Side(style="thin", color=MID_GREY),
    bottom=Side(style="thin", color=MID_GREY),
)

# ── Sheet 1: Overall Standings ──
ws = wb.active
ws.title = "Overall Standings"
ws.sheet_properties.tabColor = MASTERS_GREEN

# Title
ws.merge_cells("A1:G1")
ws["A1"] = "Golf Majors Sweepstake 2026"
ws["A1"].font = Font(name="Calibri", size=18, bold=True, color=WHITE)
ws["A1"].fill = PatternFill(start_color=MASTERS_GREEN, end_color=MASTERS_GREEN, fill_type="solid")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 40

# Subtitle
ws.merge_cells("A2:G2")
ws["A2"] = "£10 entry  •  £60 pot  •  Lowest combined score wins"
ws["A2"].font = Font(name="Calibri", size=11, color=MASTERS_GOLD)
ws["A2"].fill = PatternFill(start_color=MASTERS_GREEN, end_color=MASTERS_GREEN, fill_type="solid")
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 25

# Rules
ws.merge_cells("A3:G3")
ws["A3"] = ""
ws.row_dimensions[3].height = 8

rules = [
    "RULES: 2 Europeans + 1 American + 1 Rest of World per major",
    "No golfer can be picked more than once across all 4 majors",
    "Missed cut: R1+R2 to par + worst R3 + worst R4 from field",
    "Picks lock at the first tee shot of each major",
]
for i, rule in enumerate(rules):
    row = 4 + i
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = rule
    ws[f"A{row}"].font = Font(name="Calibri", size=10, color="666666")
    ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18

# Spacer
ws.row_dimensions[8].height = 10

# Headers
headers = ["Pos", "Player", "Masters", "PGA", "US Open", "The Open", "TOTAL"]
header_row = 9
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=header_row, column=col, value=h)
    cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    cell.fill = PatternFill(start_color=MASTERS_GREEN, end_color=MASTERS_GREEN, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
ws.row_dimensions[header_row].height = 28

# Player rows
for i, player in enumerate(PLAYERS):
    row = header_row + 1 + i
    bg = WHITE if i % 2 == 0 else LIGHT_GREY
    fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

    # Pos
    cell = ws.cell(row=row, column=1, value=i + 1)
    cell.font = Font(name="Calibri", size=12, bold=True, color=MASTERS_GREEN)
    cell.fill = fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

    # Name
    cell = ws.cell(row=row, column=2, value=player)
    cell.font = Font(name="Calibri", size=12, bold=True, color=DARK_TEXT)
    cell.fill = fill
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = thin_border

    # Major scores (empty for manual entry)
    for col in range(3, 7):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        cell.font = Font(name="Calibri", size=12, bold=True)
        cell.number_format = '0;-0;"E"'

    # Total formula
    cell = ws.cell(row=row, column=7)
    cell.value = f"=SUM(C{row}:F{row})"
    cell.font = Font(name="Calibri", size=13, bold=True, color=MASTERS_GREEN)
    cell.fill = fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    cell.number_format = '0;-0;"E"'

    ws.row_dimensions[row].height = 28

# Column widths
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 14
for c in ["C", "D", "E", "F"]:
    ws.column_dimensions[c].width = 12
ws.column_dimensions["G"].width = 10

# ── Sheets 2-5: One per major ──
for major in MAJORS:
    ws = wb.create_sheet(title=major["name"].replace("The ", ""))
    accent = MAJOR_COLOURS[major["name"]]
    ws.sheet_properties.tabColor = accent

    # Header
    ws.merge_cells("A1:F1")
    ws["A1"] = major["name"]
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, italic=True, color=WHITE)
    ws["A1"].fill = PatternFill(start_color=accent, end_color=accent, fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Venue & date
    ws.merge_cells("A2:F2")
    ws["A2"] = f"{major['venue']}  •  {major['date']}"
    ws["A2"].font = Font(name="Calibri", size=11, color=WHITE)
    ws["A2"].fill = PatternFill(start_color=accent, end_color=accent, fill_type="solid")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 25

    # Spacer
    ws.row_dimensions[3].height = 10

    # Picks section header
    ws.merge_cells("A4:F4")
    ws["A4"] = "PICKS"
    ws["A4"].font = Font(name="Calibri", size=13, bold=True, color=accent)
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[4].height = 25

    # Picks table headers
    pick_headers = ["Player", "European 1", "European 2", "American", "Rest of World"]
    for col, h in enumerate(pick_headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill(start_color=accent, end_color=accent, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[5].height = 25

    # Picks rows
    for i, player in enumerate(PLAYERS):
        row = 6 + i
        bg = WHITE if i % 2 == 0 else LIGHT_GREY
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

        cell = ws.cell(row=row, column=1, value=player)
        cell.font = Font(name="Calibri", size=11, bold=True, color=DARK_TEXT)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border

        for col in range(2, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=11)

        ws.row_dimensions[row].height = 26

    # Spacer
    spacer_row = 6 + len(PLAYERS)
    ws.row_dimensions[spacer_row].height = 15

    # Scores section header
    scores_header_row = spacer_row + 1
    ws.merge_cells(f"A{scores_header_row}:F{scores_header_row}")
    ws[f"A{scores_header_row}"] = "SCORES"
    ws[f"A{scores_header_row}"].font = Font(name="Calibri", size=13, bold=True, color=accent)
    ws[f"A{scores_header_row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[scores_header_row].height = 25

    # Scores table headers
    score_headers = ["Pos", "Player", "European 1", "European 2", "American", "RoW", "Total"]
    score_head_row = scores_header_row + 1
    for col, h in enumerate(score_headers, 1):
        cell = ws.cell(row=score_head_row, column=col, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill(start_color=accent, end_color=accent, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[score_head_row].height = 25

    # Scores rows
    for i, player in enumerate(PLAYERS):
        row = score_head_row + 1 + i
        bg = WHITE if i % 2 == 0 else LIGHT_GREY
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

        # Pos
        cell = ws.cell(row=row, column=1, value=i + 1)
        cell.font = Font(name="Calibri", size=12, bold=True, color=accent)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

        # Name
        cell = ws.cell(row=row, column=2, value=player)
        cell.font = Font(name="Calibri", size=11, bold=True, color=DARK_TEXT)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border

        # Golfer scores (empty for manual entry)
        for col in range(3, 7):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=11)
            cell.number_format = '0;-0;"E"'

        # Total formula
        cell = ws.cell(row=row, column=7)
        cell.value = f"=SUM(C{row}:F{row})"
        cell.font = Font(name="Calibri", size=12, bold=True, color=accent)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        cell.number_format = '0;-0;"E"'

        ws.row_dimensions[row].height = 26

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14
    for c in ["C", "D", "E", "F"]:
        ws.column_dimensions[c].width = 16
    ws.column_dimensions["G"].width = 10

# ── Sheet 6: All Golfers Used ──
ws = wb.create_sheet(title="Golfer Tracker")
ws.sheet_properties.tabColor = MASTERS_GOLD

ws.merge_cells("A1:E1")
ws["A1"] = "Golfer Usage Tracker"
ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=WHITE)
ws["A1"].fill = PatternFill(start_color=MASTERS_GREEN, end_color=MASTERS_GREEN, fill_type="solid")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 35

ws.merge_cells("A2:E2")
ws["A2"] = "No golfer can be picked more than once across all 4 majors"
ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="666666")
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 22

tracker_headers = ["Golfer Name", "Region", "Picked By", "Major", "Notes"]
for col, h in enumerate(tracker_headers, 1):
    cell = ws.cell(row=3, column=col, value=h)
    cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill(start_color=MASTERS_GREEN, end_color=MASTERS_GREEN, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
ws.row_dimensions[3].height = 25

# 50 empty rows for golfer tracking
for i in range(50):
    row = 4 + i
    bg = WHITE if i % 2 == 0 else LIGHT_GREY
    fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.border = thin_border
        cell.font = Font(name="Calibri", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22

ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 20

# Save
output = "/Users/andrewjames/Desktop/claude_projects/golf-sweepstake/Golf Majors Sweepstake 2026.xlsx"
wb.save(output)
print(f"Saved: {output}")
